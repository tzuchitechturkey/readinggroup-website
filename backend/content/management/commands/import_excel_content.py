"""
Management command: import_excel_content
========================================
Imports all content from the Excel entry template into the database.

Usage
-----
    python manage.py import_excel_content --file /path/to/Content_Entry_Template.xlsx

Options
-------
    --file       Path to the Excel template (required)
    --dry-run    Validate and preview without saving anything
    --skip-images  Skip downloading images (for ImageField-only columns)
    --reset      Delete all existing content before importing (⚠ destructive)
"""

from __future__ import annotations

import io
import os
import urllib.request
import urllib.error
from datetime import datetime, date, time
from typing import Any, Optional

from django.core.files.base import ContentFile
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify

try:
    import openpyxl
except ImportError:
    raise CommandError("openpyxl is required: pip install openpyxl")

from content.enums import VideoType, LearnType, LearnCategoryDirection
from content.models import (
    Video,
    VideoCategory,
    Learn,
    LearnCategory,
    RelatedReports,
    RelatedReportsCategory,
    PhotoCollection,
    Photo,
    LatestNews,
    LatestNewsImage,
    EventCommunity,
    OurTeam,
    OurTeamImage,
    Book,
    BookReview,
    HistoryEvent,
    HistoryEventImage,
    SocialMedia,
)

# ── Helpers ────────────────────────────────────────────────────────────────────

DATA_START_ROW = 4  # rows 1-3 are headers/type/notes


def _val(ws, row: int, col: int) -> Optional[str]:
    """Return a stripped string value from a cell, or None if blank."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _bool(raw: Optional[str], default: bool = False) -> bool:
    if raw is None:
        return default
    return raw.strip().upper() == "TRUE"


def _int(raw: Optional[str], default: int = 0) -> int:
    if raw is None:
        return default
    try:
        return int(float(raw))
    except (ValueError, TypeError):
        return default


def _datetime(raw: Optional[str]) -> Optional[datetime]:
    if not raw:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(raw, fmt)
            return timezone.make_aware(dt)
        except ValueError:
            continue
    return None


def _date(raw: Optional[str]) -> Optional[date]:
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _time_val(raw: Optional[str]) -> Optional[time]:
    if not raw:
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    return None


def _json_list(*items) -> list:
    """Build a JSON list from multiple optional string values."""
    return [i for i in items if i]


def _gdrive_direct(url: str) -> str:
    """Convert a Google Drive share URL to a direct download URL."""
    if "drive.google.com/file/d/" in url:
        try:
            file_id = url.split("/file/d/")[1].split("/")[0]
            return f"https://drive.google.com/uc?export=download&id={file_id}"
        except IndexError:
            pass
    return url


def _download_image(url: str, filename: str) -> Optional[ContentFile]:
    """Download an image from a URL and return a ContentFile, or None on failure."""
    try:
        direct_url = _gdrive_direct(url)
        req = urllib.request.Request(
            direct_url,
            headers={"User-Agent": "Mozilla/5.0 (compatible; ImportScript/1.0)"},
        )
        with urllib.request.urlopen(req, timeout=15) as response:
            data = response.read()
        return ContentFile(data, name=filename)
    except (urllib.error.URLError, Exception):
        return None


def _rows(ws):
    """Iterate data rows (row 4 onwards), skip completely empty rows."""
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row_vals = [
            ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)
        ]
        if any(v is not None and str(v).strip() for v in row_vals):
            yield row_idx


# ── Importers ─────────────────────────────────────────────────────────────────


class Importer:
    def __init__(self, wb, stdout, dry_run: bool, skip_images: bool):
        self.wb = wb
        self.stdout = stdout
        self.dry_run = dry_run
        self.skip_images = skip_images
        self.summary: dict[str, dict] = {}
        self.errors: list[str] = []

    def _log(self, msg: str):
        self.stdout.write(msg)

    def _sheet(self, name: str):
        if name not in self.wb.sheetnames:
            raise CommandError(f"Sheet '{name}' not found in the Excel file.")
        return self.wb[name]

    def _record(
        self, sheet: str, created: int, updated: int, skipped: int, errors: int
    ):
        self.summary[sheet] = dict(
            created=created, updated=updated, skipped=skipped, errors=errors
        )

    # ── Video Categories ──────────────────────────────────────────────────────
    def import_video_categories(self):
        ws = self._sheet("VideoCategories")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            name = _val(ws, row, 1)
            if not name:
                skipped += 1
                continue
            try:
                data = dict(
                    description=_val(ws, row, 2) or "",
                    is_active=_bool(_val(ws, row, 3), default=True),
                    order=_int(_val(ws, row, 4), default=0),
                )
                if not self.dry_run:
                    obj, was_created = VideoCategory.objects.update_or_create(
                        name=name, defaults=data
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"VideoCategories row {row}: {e}")
                errors += 1
        self._record("VideoCategories", created, updated, skipped, errors)
        self._log(
            f"  VideoCategories   → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Videos ───────────────────────────────────────────────────────────────
    def import_videos(self):
        ws = self._sheet("Videos")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            title = _val(ws, row, 1)
            video_url = _val(ws, row, 2)
            if not title or not video_url:
                skipped += 1
                continue
            try:
                cat_name = _val(ws, row, 5)
                category = (
                    VideoCategory.objects.filter(name=cat_name).first()
                    if cat_name
                    else None
                )
                video_type = _val(ws, row, 3)
                if video_type and video_type not in VideoType.values:
                    video_type = None

                thumb_urls = _json_list(
                    _val(ws, row, 10),
                    _val(ws, row, 11),
                    _val(ws, row, 12),
                )
                guest_speakers = _json_list(
                    _val(ws, row, 13),
                    _val(ws, row, 14),
                    _val(ws, row, 15),
                )
                data = dict(
                    video_url=video_url,
                    video_type=video_type or "",
                    language=_val(ws, row, 4) or "",
                    category=category,
                    description=_val(ws, row, 6) or "",
                    duration=_val(ws, row, 7),
                    happened_at=_datetime(_val(ws, row, 8)),
                    reference_code=_val(ws, row, 9) or "",
                    thumbnail_url=thumb_urls,
                    guest_speakers=guest_speakers,
                )
                if not self.dry_run:
                    obj, was_created = Video.objects.update_or_create(
                        title=title, defaults=data
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"Videos row {row} ('{title}'): {e}")
                errors += 1
        self._record("Videos", created, updated, skipped, errors)
        self._log(
            f"  Videos            → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Learn Categories ──────────────────────────────────────────────────────
    def import_learn_categories(self):
        ws = self._sheet("LearnCategories")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            name = _val(ws, row, 1)
            learn_type = _val(ws, row, 2)
            if not name or not learn_type:
                skipped += 1
                continue
            if learn_type not in LearnType.values:
                self.errors.append(
                    f"LearnCategories row {row}: invalid learn_type '{learn_type}' – must be 'cards' or 'posters'"
                )
                errors += 1
                continue
            try:
                direction = _val(ws, row, 4) or ""
                if direction and direction not in LearnCategoryDirection.values:
                    direction = ""
                data = dict(
                    learn_type=learn_type,
                    description=_val(ws, row, 3) or "",
                    direction=direction,
                    is_active=_bool(_val(ws, row, 5), default=True),
                    order=_int(_val(ws, row, 6), default=0),
                )
                if not self.dry_run:
                    obj, was_created = LearnCategory.objects.update_or_create(
                        name=name, learn_type=learn_type, defaults=data
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"LearnCategories row {row} ('{name}'): {e}")
                errors += 1
        self._record("LearnCategories", created, updated, skipped, errors)
        self._log(
            f"  LearnCategories   → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Learn ─────────────────────────────────────────────────────────────────
    def import_learn(self):
        ws = self._sheet("Learn")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            title = _val(ws, row, 1)
            cat_name = _val(ws, row, 2)
            if not title or not cat_name:
                skipped += 1
                continue
            try:
                category = LearnCategory.objects.filter(name=cat_name).first()
                if not category:
                    self.errors.append(
                        f"Learn row {row}: category '{cat_name}' not found – create it in LearnCategories first"
                    )
                    errors += 1
                    continue
                data = dict(
                    category=category,
                    subtitle=_val(ws, row, 3) or "",
                    image_url=_val(ws, row, 4) or "",
                    happened_at=_datetime(_val(ws, row, 5)),
                )
                if not self.dry_run:
                    obj, was_created = Learn.objects.update_or_create(
                        title=title, defaults=data
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"Learn row {row} ('{title}'): {e}")
                errors += 1
        self._record("Learn", created, updated, skipped, errors)
        self._log(
            f"  Learn             → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Related Reports Categories ────────────────────────────────────────────
    def import_related_reports_categories(self):
        ws = self._sheet("RelatedReportsCategories")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            title = _val(ws, row, 1)
            if not title:
                skipped += 1
                continue
            try:
                data = dict(
                    description=_val(ws, row, 2) or "",
                    is_active=_bool(_val(ws, row, 3), default=True),
                )
                if not self.dry_run:
                    obj, was_created = RelatedReportsCategory.objects.update_or_create(
                        title=title, defaults=data
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"RelatedReportsCategories row {row}: {e}")
                errors += 1
        self._record("RelatedReportsCategories", created, updated, skipped, errors)
        self._log(
            f"  RelatedReportsCats→ created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Related Reports ───────────────────────────────────────────────────────
    def import_related_reports(self):
        ws = self._sheet("RelatedReports")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            title = _val(ws, row, 1)
            if not title:
                skipped += 1
                continue
            try:
                cat_title = _val(ws, row, 2)
                category = (
                    RelatedReportsCategory.objects.filter(title=cat_title).first()
                    if cat_title
                    else None
                )
                thumb_urls = _json_list(_val(ws, row, 7), _val(ws, row, 8))
                data = dict(
                    category=category,
                    description=_val(ws, row, 3) or "",
                    external_link=_val(ws, row, 4),
                    happened_at=_datetime(_val(ws, row, 5)),
                    duration=_val(ws, row, 6),
                    thumbnail_url=thumb_urls,
                )
                if not self.dry_run:
                    obj, was_created = RelatedReports.objects.update_or_create(
                        title=title, defaults=data
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"RelatedReports row {row} ('{title}'): {e}")
                errors += 1
        self._record("RelatedReports", created, updated, skipped, errors)
        self._log(
            f"  RelatedReports    → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Photo Collections ─────────────────────────────────────────────────────
    def import_photo_collections(self):
        ws = self._sheet("PhotoCollections")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            title = _val(ws, row, 1)
            if not title:
                skipped += 1
                continue
            try:
                cover_url = _val(ws, row, 3)
                happened_at = _datetime(_val(ws, row, 4))
                description = _val(ws, row, 2) or ""
                if not self.dry_run:
                    obj, was_created = PhotoCollection.objects.update_or_create(
                        title=title,
                        defaults=dict(description=description, happened_at=happened_at),
                    )
                    # Download cover image if provided
                    if cover_url and not self.skip_images:
                        ext = os.path.splitext(cover_url.split("?")[0])[1] or ".jpg"
                        fname = f"cover_{slugify(title)}{ext}"
                        cf = _download_image(cover_url, fname)
                        if cf:
                            obj.image.save(fname, cf, save=True)
                        else:
                            self.errors.append(
                                f"PhotoCollections row {row}: could not download cover image from {cover_url}"
                            )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"PhotoCollections row {row} ('{title}'): {e}")
                errors += 1
        self._record("PhotoCollections", created, updated, skipped, errors)
        self._log(
            f"  PhotoCollections  → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Photos ────────────────────────────────────────────────────────────────
    def import_photos(self):
        ws = self._sheet("Photos")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            col_title = _val(ws, row, 1)
            image_url = _val(ws, row, 2)
            order_raw = _val(ws, row, 4)
            if not col_title or not image_url:
                skipped += 1
                continue
            try:
                collection = PhotoCollection.objects.filter(title=col_title).first()
                if not collection:
                    self.errors.append(
                        f"Photos row {row}: collection '{col_title}' not found"
                    )
                    errors += 1
                    continue
                order = _int(order_raw, default=1)
                caption = _val(ws, row, 3) or ""
                if not self.dry_run:
                    obj, was_created = Photo.objects.update_or_create(
                        collection=collection,
                        order=order,
                        defaults=dict(caption=caption),
                    )
                    if not self.skip_images:
                        ext = os.path.splitext(image_url.split("?")[0])[1] or ".jpg"
                        fname = f"photo_{slugify(col_title)}_{order}{ext}"
                        cf = _download_image(image_url, fname)
                        if cf:
                            obj.image.save(fname, cf, save=True)
                        else:
                            self.errors.append(
                                f"Photos row {row}: could not download image from {image_url}"
                            )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"Photos row {row}: {e}")
                errors += 1
        self._record("Photos", created, updated, skipped, errors)
        self._log(
            f"  Photos            → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Latest News ───────────────────────────────────────────────────────────
    def import_latest_news(self):
        ws = self._sheet("LatestNews")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            title = _val(ws, row, 1)
            if not title:
                skipped += 1
                continue
            try:
                data = dict(
                    description=_val(ws, row, 2) or "",
                    happened_at=_datetime(_val(ws, row, 3)),
                    is_new=_bool(_val(ws, row, 4), default=False),
                )
                if not self.dry_run:
                    obj, was_created = LatestNews.objects.update_or_create(
                        title=title, defaults=data
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"LatestNews row {row} ('{title}'): {e}")
                errors += 1
        self._record("LatestNews", created, updated, skipped, errors)
        self._log(
            f"  LatestNews        → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Latest News Images ────────────────────────────────────────────────────
    def import_latest_news_images(self):
        ws = self._sheet("LatestNewsImages")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            news_title = _val(ws, row, 1)
            image_url = _val(ws, row, 2)
            order_raw = _val(ws, row, 4)
            if not news_title or not image_url:
                skipped += 1
                continue
            try:
                news = LatestNews.objects.filter(title=news_title).first()
                if not news:
                    self.errors.append(
                        f"LatestNewsImages row {row}: news '{news_title}' not found"
                    )
                    errors += 1
                    continue
                order = _int(order_raw, default=1)
                caption = _val(ws, row, 3) or ""
                if not self.dry_run:
                    obj, was_created = LatestNewsImage.objects.update_or_create(
                        latest_news=news,
                        order=order,
                        defaults=dict(caption=caption),
                    )
                    if not self.skip_images:
                        ext = os.path.splitext(image_url.split("?")[0])[1] or ".jpg"
                        fname = f"news_{slugify(news_title)}_{order}{ext}"
                        cf = _download_image(image_url, fname)
                        if cf:
                            obj.image.save(fname, cf, save=True)
                        else:
                            self.errors.append(
                                f"LatestNewsImages row {row}: could not download image from {image_url}"
                            )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"LatestNewsImages row {row}: {e}")
                errors += 1
        self._record("LatestNewsImages", created, updated, skipped, errors)
        self._log(
            f"  LatestNewsImages  → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Event Community ───────────────────────────────────────────────────────
    def import_event_community(self):
        ws = self._sheet("EventCommunity")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            title = _val(ws, row, 1)
            if not title:
                skipped += 1
                continue
            try:
                learn_title = _val(ws, row, 6)
                learn = (
                    Learn.objects.filter(
                        title=learn_title, category__learn_type=LearnType.POSTERS
                    ).first()
                    if learn_title
                    else None
                )
                guest_speakers = _json_list(
                    _val(ws, row, 7),
                    _val(ws, row, 8),
                    _val(ws, row, 9),
                )
                data = dict(
                    start_event_date=_date(_val(ws, row, 2)),
                    start_event_time=_time_val(_val(ws, row, 3)),
                    duration=_val(ws, row, 4),
                    live_stream_link=_val(ws, row, 5),
                    learn=learn,
                    guest_speakers=guest_speakers,
                )
                if not self.dry_run:
                    obj, was_created = EventCommunity.objects.update_or_create(
                        title=title, defaults=data
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"EventCommunity row {row} ('{title}'): {e}")
                errors += 1
        self._record("EventCommunity", created, updated, skipped, errors)
        self._log(
            f"  EventCommunity    → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Our Team ──────────────────────────────────────────────────────────────
    def import_our_team(self):
        ws = self._sheet("OurTeam")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            title = _val(ws, row, 1)
            if not title:
                skipped += 1
                continue
            try:
                image_urls = _json_list(_val(ws, row, 4), _val(ws, row, 5))
                data = dict(
                    description=_val(ws, row, 2) or "",
                    is_heart=_bool(_val(ws, row, 3), default=False),
                    image_url=image_urls,
                )
                if not self.dry_run:
                    obj, was_created = OurTeam.objects.update_or_create(
                        title=title, defaults=data
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"OurTeam row {row} ('{title}'): {e}")
                errors += 1
        self._record("OurTeam", created, updated, skipped, errors)
        self._log(
            f"  OurTeam           → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Our Team Images ───────────────────────────────────────────────────────
    def import_our_team_images(self):
        ws = self._sheet("OurTeamImages")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            member_title = _val(ws, row, 1)
            image_url = _val(ws, row, 2)
            order_raw = _val(ws, row, 4)
            if not member_title or not image_url:
                skipped += 1
                continue
            try:
                member = OurTeam.objects.filter(title=member_title).first()
                if not member:
                    self.errors.append(
                        f"OurTeamImages row {row}: team member '{member_title}' not found"
                    )
                    errors += 1
                    continue
                order = _int(order_raw, default=1)
                caption = _val(ws, row, 3) or ""
                if not self.dry_run:
                    obj, was_created = OurTeamImage.objects.update_or_create(
                        our_team=member,
                        order=order,
                        defaults=dict(caption=caption),
                    )
                    if not self.skip_images:
                        ext = os.path.splitext(image_url.split("?")[0])[1] or ".jpg"
                        fname = f"team_{slugify(member_title)}_{order}{ext}"
                        cf = _download_image(image_url, fname)
                        if cf:
                            obj.image.save(fname, cf, save=True)
                        else:
                            self.errors.append(
                                f"OurTeamImages row {row}: could not download image from {image_url}"
                            )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"OurTeamImages row {row}: {e}")
                errors += 1
        self._record("OurTeamImages", created, updated, skipped, errors)
        self._log(
            f"  OurTeamImages     → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Books ─────────────────────────────────────────────────────────────────
    def import_books(self):
        ws = self._sheet("Books")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            title = _val(ws, row, 1)
            if not title:
                skipped += 1
                continue
            try:
                image_url = _val(ws, row, 3)
                cover_image_url = _val(ws, row, 4)
                if not self.dry_run:
                    obj, was_created = Book.objects.update_or_create(
                        title=title,
                        defaults=dict(description=_val(ws, row, 2) or ""),
                    )
                    if not self.skip_images:
                        if image_url:
                            ext = os.path.splitext(image_url.split("?")[0])[1] or ".jpg"
                            fname = f"book_{slugify(title)}{ext}"
                            cf = _download_image(image_url, fname)
                            if cf:
                                obj.image.save(fname, cf, save=True)
                            else:
                                self.errors.append(
                                    f"Books row {row}: could not download image from {image_url}"
                                )
                        if cover_image_url:
                            ext = (
                                os.path.splitext(cover_image_url.split("?")[0])[1]
                                or ".jpg"
                            )
                            fname = f"book_cover_{slugify(title)}{ext}"
                            cf = _download_image(cover_image_url, fname)
                            if cf:
                                obj.cover_image.save(fname, cf, save=True)
                            else:
                                self.errors.append(
                                    f"Books row {row}: could not download cover from {cover_image_url}"
                                )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"Books row {row} ('{title}'): {e}")
                errors += 1
        self._record("Books", created, updated, skipped, errors)
        self._log(
            f"  Books             → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Book Reviews ──────────────────────────────────────────────────────────
    def import_book_reviews(self):
        ws = self._sheet("BookReviews")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            book_title = _val(ws, row, 1)
            image_url = _val(ws, row, 2)
            order_raw = _val(ws, row, 3)
            if not book_title or not image_url:
                skipped += 1
                continue
            try:
                book = Book.objects.filter(title=book_title).first()
                if not book:
                    self.errors.append(
                        f"BookReviews row {row}: book '{book_title}' not found"
                    )
                    errors += 1
                    continue
                order = _int(order_raw, default=1)
                if not self.dry_run:
                    obj, was_created = BookReview.objects.update_or_create(
                        book=book, order=order, defaults={}
                    )
                    if not self.skip_images:
                        ext = os.path.splitext(image_url.split("?")[0])[1] or ".jpg"
                        fname = f"book_review_{slugify(book_title)}_{order}{ext}"
                        cf = _download_image(image_url, fname)
                        if cf:
                            obj.image.save(fname, cf, save=True)
                        else:
                            self.errors.append(
                                f"BookReviews row {row}: could not download image from {image_url}"
                            )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"BookReviews row {row}: {e}")
                errors += 1
        self._record("BookReviews", created, updated, skipped, errors)
        self._log(
            f"  BookReviews       → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── History Events ────────────────────────────────────────────────────────
    def import_history_events(self):
        ws = self._sheet("HistoryEvents")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            year_raw = _val(ws, row, 1)
            title = _val(ws, row, 3)
            if not year_raw or not title:
                skipped += 1
                continue
            try:
                year = _int(year_raw)
                month = _int(_val(ws, row, 2), default=1)
                data = dict(
                    month=month,
                    sub_title_one=_val(ws, row, 4) or "",
                    sub_title_two=_val(ws, row, 5) or "",
                    description=_val(ws, row, 6) or "",
                )
                if not self.dry_run:
                    obj, was_created = HistoryEvent.objects.update_or_create(
                        year=year, title=title, defaults=data
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"HistoryEvents row {row} ('{title}'): {e}")
                errors += 1
        self._record("HistoryEvents", created, updated, skipped, errors)
        self._log(
            f"  HistoryEvents     → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── History Event Images ──────────────────────────────────────────────────
    def import_history_event_images(self):
        ws = self._sheet("HistoryEventImages")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            event_year = _val(ws, row, 1)
            event_title = _val(ws, row, 2)
            image_url = _val(ws, row, 3)
            caption = _val(ws, row, 4)
            order_raw = _val(ws, row, 5)
            if not event_year or not event_title or not image_url:
                skipped += 1
                continue
            try:
                event = HistoryEvent.objects.filter(
                    year=_int(event_year), title=event_title
                ).first()
                if not event:
                    self.errors.append(
                        f"HistoryEventImages row {row}: event '{event_title}' ({event_year}) not found"
                    )
                    errors += 1
                    continue
                order = _int(order_raw, default=1)
                if not self.dry_run:
                    obj, was_created = HistoryEventImage.objects.update_or_create(
                        event=event,
                        order=order,
                        defaults=dict(caption=caption or ""),
                    )
                    if not self.skip_images:
                        ext = os.path.splitext(image_url.split("?")[0])[1] or ".jpg"
                        fname = f"history_{slugify(event_title)}_{order}{ext}"
                        cf = _download_image(image_url, fname)
                        if cf:
                            obj.image.save(fname, cf, save=True)
                        else:
                            self.errors.append(
                                f"HistoryEventImages row {row}: could not download image from {image_url}"
                            )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"HistoryEventImages row {row}: {e}")
                errors += 1
        self._record("HistoryEventImages", created, updated, skipped, errors)
        self._log(
            f"  HistoryEventImages→ created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Social Media ──────────────────────────────────────────────────────────
    def import_social_media(self):
        ws = self._sheet("SocialMedia")
        created = updated = skipped = errors = 0
        for row in _rows(ws):
            platform = _val(ws, row, 1)
            url = _val(ws, row, 2)
            if not platform or not url:
                skipped += 1
                continue
            try:
                if not self.dry_run:
                    obj, was_created = SocialMedia.objects.update_or_create(
                        platform=platform, defaults=dict(url=url)
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1
                else:
                    created += 1
            except Exception as e:
                self.errors.append(f"SocialMedia row {row} ('{platform}'): {e}")
                errors += 1
        self._record("SocialMedia", created, updated, skipped, errors)
        self._log(
            f"  SocialMedia       → created={created} updated={updated} skipped={skipped} errors={errors}"
        )

    # ── Run all ───────────────────────────────────────────────────────────────
    def run(self):
        steps = [
            # Categories first (required before FK lookups)
            self.import_video_categories,
            self.import_learn_categories,
            self.import_related_reports_categories,
            # Main content
            self.import_videos,
            self.import_learn,
            self.import_related_reports,
            # Collections → children
            self.import_photo_collections,
            self.import_photos,
            self.import_latest_news,
            self.import_latest_news_images,
            self.import_event_community,
            self.import_our_team,
            self.import_our_team_images,
            self.import_books,
            self.import_book_reviews,
            self.import_history_events,
            self.import_history_event_images,
            self.import_social_media,
        ]
        for step in steps:
            step()


# ── Management command ────────────────────────────────────────────────────────


class Command(BaseCommand):
    help = "Import content from the Excel entry template into the database."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            "-f",
            required=True,
            help="Path to the Excel template file (Content_Entry_Template.xlsx)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            default=False,
            help="Validate and count rows without saving anything to the database",
        )
        parser.add_argument(
            "--skip-images",
            action="store_true",
            default=False,
            help="Skip downloading images (faster; useful for testing text content first)",
        )
        parser.add_argument(
            "--reset",
            action="store_true",
            default=False,
            help="⚠ Delete ALL existing content before importing (irreversible)",
        )

    def handle(self, *args, **options):
        file_path = options["file"]
        dry_run = options["dry_run"]
        skip_images = options["skip_images"]
        reset = options["reset"]

        if not os.path.exists(file_path):
            raise CommandError(f"File not found: {file_path}")

        self.stdout.write(
            self.style.MIGRATE_HEADING("\n═══════════════════════════════════════")
        )
        self.stdout.write(self.style.MIGRATE_HEADING(" Content Excel Importer"))
        self.stdout.write(
            self.style.MIGRATE_HEADING("═══════════════════════════════════════")
        )
        self.stdout.write(f" File       : {file_path}")
        self.stdout.write(
            f" Dry run    : {'YES – nothing will be saved' if dry_run else 'NO – data will be saved'}"
        )
        self.stdout.write(f" Skip images: {skip_images}")
        self.stdout.write(f" Reset      : {reset}")
        self.stdout.write("")

        if reset and not dry_run:
            self.stdout.write(
                self.style.WARNING(
                    "⚠  --reset flag detected: deleting all existing content…"
                )
            )
            models_to_clear = [
                HistoryEventImage,
                HistoryEvent,
                BookReview,
                Book,
                OurTeamImage,
                OurTeam,
                EventCommunity,
                LatestNewsImage,
                LatestNews,
                Photo,
                PhotoCollection,
                RelatedReports,
                RelatedReportsCategory,
                Learn,
                LearnCategory,
                Video,
                VideoCategory,
                SocialMedia,
            ]
            for model in models_to_clear:
                count, _ = model.objects.all().delete()
                self.stdout.write(f"   Deleted {count:>5} {model.__name__} records")
            self.stdout.write("")

        self.stdout.write("Loading workbook…")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise CommandError(f"Could not open Excel file: {e}")

        self.stdout.write("Importing sheets:\n")

        importer = Importer(wb, self.stdout, dry_run=dry_run, skip_images=skip_images)

        if dry_run:
            importer.run()
        else:
            with transaction.atomic():
                importer.run()

        # ── Summary ───────────────────────────────────────────────────────────
        self.stdout.write("")
        self.stdout.write(
            self.style.MIGRATE_HEADING("═══ Summary ════════════════════════════")
        )
        total_created = total_updated = total_skipped = total_errors = 0
        for sheet, counts in importer.summary.items():
            total_created += counts["created"]
            total_updated += counts["updated"]
            total_skipped += counts["skipped"]
            total_errors += counts["errors"]
        self.stdout.write(f" Total created : {total_created}")
        self.stdout.write(f" Total updated : {total_updated}")
        self.stdout.write(f" Total skipped : {total_skipped}")
        self.stdout.write(f" Total errors  : {total_errors}")
        self.stdout.write("")

        if importer.errors:
            self.stdout.write(
                self.style.WARNING(f"⚠  {len(importer.errors)} issue(s) encountered:")
            )
            for err in importer.errors:
                self.stdout.write(self.style.ERROR(f"   • {err}"))
            self.stdout.write("")

        if dry_run:
            self.stdout.write(
                self.style.WARNING("DRY RUN complete – no data was saved.")
            )
            self.stdout.write("Re-run without --dry-run to commit the import.")
        else:
            if total_errors == 0:
                self.stdout.write(
                    self.style.SUCCESS("✓ Import complete with no errors.")
                )
            else:
                self.stdout.write(
                    self.style.WARNING(
                        f"Import complete with {total_errors} error(s). Check the list above."
                    )
                )
        self.stdout.write("")
